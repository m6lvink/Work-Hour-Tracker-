from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
import os
import logging
import json
from datetime import datetime
import re
from dotenv import load_dotenv
import pandas as pd
import openpyxl
from io import BytesIO

from google_auth_oauthlib.flow import Flow

# Initialize the OAuth flow from the Render-mounted secret file
flow = Flow.from_client_secrets_file(
    os.environ["GOOGLE_CREDENTIALS_FILE"],
    scopes=[
        "https://www.googleapis.com/auth/userinfo.email",
        "https://www.googleapis.com/auth/userinfo.profile"
    ],
    redirect_uri="https://<YOUR-RENDER-URL>/oauth2callback"
)



# Load environment variables
load_dotenv()



# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', os.urandom(24))

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///workhours.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

def init_db():
    """Initialize the database and handle migrations."""
    try:
        with app.app_context():
            # Check if database exists
            db_exists = os.path.exists('workhours.db')
            logger.info(f"Database exists: {db_exists}")

            # Create all tables
            db.create_all()
            logger.info("Database tables created successfully")

            # Check if we need to migrate task to description
            if db_exists:
                try:
                    # Check if task column exists
                    result = db.session.execute("PRAGMA table_info(entry)")
                    columns = [row[1] for row in result]
                    
                    if 'task' in columns and 'description' not in columns:
                        logger.info("Migrating task column to description")
                        # Add description column
                        db.session.execute("ALTER TABLE entry ADD COLUMN description TEXT")
                        # Copy data from task to description
                        db.session.execute("UPDATE entry SET description = task")
                        # Drop task column
                        db.session.execute("ALTER TABLE entry DROP COLUMN task")
                        db.session.commit()
                        logger.info("Migration completed successfully")
                except Exception as e:
                    logger.error(f"Error during migration: {str(e)}", exc_info=True)
                    db.session.rollback()

    except Exception as e:
        logger.error(f"Error initializing database: {str(e)}", exc_info=True)
        raise

# Initialize database
init_db()

def export_to_excel():
    """Export entries to a nicely formatted Excel worksheet."""
    try:
        # Get entries for the current user
        entries = Entry.query.filter_by(user_id=session['user_id']).order_by(Entry.date).all()
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Work Sheet"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        
        # Add title
        ws['A1'] = "Work Sheet"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Add headers
        headers = ['Day', 'Hours', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Add data
        for row, entry in enumerate(entries, 3):
            ws.cell(row=row, column=1).value = entry.date
            ws.cell(row=row, column=2).value = entry.hours
            ws.cell(row=row, column=3).value = entry.description
        
        # Add borders
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=len(entries)+2, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        # Create timestamp for filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'workhours_{timestamp}.xlsx'
        
        # Save the workbook
        wb.save(excel_filename)
        logger.info(f"Successfully exported data to {excel_filename}")
        return excel_filename
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return None

@app.route('/export')
def export():
    if 'user_id' not in session:
        flash('Please log in to export entries', 'warning')
        return redirect(url_for('login'))
    
    try:
        entries = Entry.query.filter_by(user_id=session['user_id']).order_by(Entry.date.desc()).all()
        
        # Create DataFrame
        data = {
            'Date': [entry.date.strftime('%Y-%m-%d') for entry in entries],
            'Hours': [entry.hours for entry in entries],
            'Description': [entry.description for entry in entries]
        }
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Work Hours')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Work Hours']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
        
        output.seek(0)
        
        # Generate filename with current date
        filename = f'work_hours_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting entries: {str(e)}", exc_info=True)
        flash('An error occurred while exporting entries', 'error')
        return redirect(url_for('dashboard'))

@app.route('/import', methods=['POST'])
def import_excel():
    if 'user_id' not in session:
        flash('Please log in to import entries', 'warning')
        return redirect(url_for('login'))
    
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('dashboard'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        df = pd.read_excel(file)
        required_columns = ['Date', 'Hours', 'Description']
        
        if not all(col in df.columns for col in required_columns):
            flash('Excel file must contain Date, Hours, and Description columns', 'error')
            return redirect(url_for('dashboard'))
        
        entries_added = 0
        for _, row in df.iterrows():
            try:
                date = pd.to_datetime(row['Date']).date()
                hours = float(row['Hours'])
                description = str(row['Description']).strip()
                
                if hours <= 0 or hours > 24:
                    continue
                
                entry = Entry(
                    user_id=session['user_id'],
                    date=date,
                    hours=hours,
                    description=description
                )
                db.session.add(entry)
                entries_added += 1
                
            except (ValueError, TypeError):
                continue
        
        db.session.commit()
        flash(f'Successfully imported {entries_added} entries!', 'success')
        
    except Exception as e:
        logger.error(f"Error importing Excel file: {str(e)}", exc_info=True)
        db.session.rollback()
        flash('An error occurred while importing the file', 'error')
    
    return redirect(url_for('dashboard'))

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(128), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    settings = db.relationship('UserSettings', backref='user', uselist=False)

class UserSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    full_name = db.Column(db.String(100), nullable=True)
    hourly_rate = db.Column(db.Float, nullable=True)
    tax_rate = db.Column(db.Float, default=0.0)  # Hawaii tax rate
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Entry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    hours = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.strftime('%Y-%m-%d'),
            'hours': self.hours,
            'description': self.description
        }

def validate_date(date_str):
    """Validate date format (YYYY-MM-DD)."""
    pattern = r'^\d{4}-\d{2}-\d{2}$'
    if not re.match(pattern, date_str):
        return False
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        return False

@app.route('/')
def index():
    return redirect(url_for('welcome'))

@app.route('/welcome')
def welcome():
    return render_template('welcome.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Please log in to access the dashboard', 'warning')
        return redirect(url_for('login'))
    
    try:
        logger.debug(f"Accessing dashboard for user_id: {session['user_id']}")
        
        # Get user
        user = User.query.get(session['user_id'])
        if not user:
            logger.error(f"User not found for user_id: {session['user_id']}")
            session.clear()  # Clear invalid session
            flash('Session expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        
        # Get entries with error handling
        try:
            entries = Entry.query.filter_by(user_id=session['user_id']).order_by(Entry.date.desc()).all()
            logger.debug(f"Found {len(entries)} entries for user")
        except Exception as e:
            logger.error(f"Error fetching entries: {str(e)}", exc_info=True)
            entries = []
        
        # Calculate total hours
        try:
            total_hours = sum(entry.hours for entry in entries)
            logger.debug(f"Total hours: {total_hours}")
        except Exception as e:
            logger.error(f"Error calculating total hours: {str(e)}", exc_info=True)
            total_hours = 0
        
        # Calculate entries this month
        try:
            current_month = datetime.now().date().replace(day=1)
            entries_this_month = len([e for e in entries if e.date >= current_month])
            logger.debug(f"Entries this month: {entries_this_month}")
        except Exception as e:
            logger.error(f"Error calculating entries this month: {str(e)}", exc_info=True)
            entries_this_month = 0
        
        # Calculate average hours per day
        try:
            avg_hours_per_day = total_hours / len(entries) if entries else 0
            logger.debug(f"Average hours per day: {avg_hours_per_day}")
        except Exception as e:
            logger.error(f"Error calculating average hours: {str(e)}", exc_info=True)
            avg_hours_per_day = 0
        
        # Calculate earnings
        total_earnings = 0
        net_earnings = 0
        try:
            if user.settings and user.settings.hourly_rate:
                total_earnings = total_hours * user.settings.hourly_rate
                if user.settings.tax_rate:
                    tax_amount = total_earnings * (user.settings.tax_rate / 100)
                    net_earnings = total_earnings - tax_amount
                logger.debug(f"Total earnings: {total_earnings}, Net earnings: {net_earnings}")
        except Exception as e:
            logger.error(f"Error calculating earnings: {str(e)}", exc_info=True)
        
        # Prepare template data
        template_data = {
            'entries': [entry.to_dict() for entry in entries],
            'total_hours': round(total_hours, 2),
            'entries_this_month': entries_this_month,
            'avg_hours_per_day': round(avg_hours_per_day, 2),
            'total_earnings': round(total_earnings, 2),
            'net_earnings': round(net_earnings, 2),
            'settings': user.settings
        }
        logger.debug(f"Template data prepared: {template_data}")
        
        return render_template('index.html', **template_data)
        
    except Exception as e:
        logger.error(f"Error accessing dashboard: {str(e)}", exc_info=True)
        flash('An error occurred while loading the dashboard', 'error')
        return redirect(url_for('welcome'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            username = request.form['username'].strip()
            password = request.form['password']
            confirm = request.form['confirm']
            email = request.form.get('email', '').strip()

            logger.debug(f"Registration attempt for username: {username}")

            # Enhanced validation
            if not username or not password:
                flash('Username and password are required', 'error')
                return render_template('register.html')

            if len(username) < 3 or len(username) > 20:
                flash('Username must be between 3 and 20 characters', 'error')
                return render_template('register.html')

            if User.query.filter_by(username=username).first():
                flash('Username already taken', 'error')
                return render_template('register.html')

            if len(password) < 8:
                flash('Password must be at least 8 characters long', 'error')
                return render_template('register.html')

            if password.lower() in ["password", "12345678", "qwerty", username.lower()]:
                flash('Password is too common or predictable', 'error')
                return render_template('register.html')

            if password != confirm:
                flash('Passwords do not match.', 'error')
                return render_template('register.html')

            if email and not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                flash('Invalid email format', 'error')
                return render_template('register.html')

            hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
            user = User(username=username, password=hashed_pw, email=email)
            db.session.add(user)
            db.session.commit()
            
            logger.info(f"Successfully registered user: {username}")
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))

        except Exception as e:
            logger.error(f"Registration error: {str(e)}")
            db.session.rollback()
            # Only show a generic error if it's not a password mismatch
            if 'Passwords do not match.' not in str(e):
                flash('An error occurred during registration', 'error')
            return render_template('register.html')

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            username = request.form['username'].strip()
            password = request.form['password']
            
            logger.debug(f"Login attempt for username: {username}")
            
            if not username or not password:
                flash('Username and password are required', 'error')
                return render_template('login.html')

            user = User.query.filter_by(username=username).first()
            if user and bcrypt.check_password_hash(user.password, password):
                session['user_id'] = user.id
                session.permanent = True  # Make session persistent
                logger.info(f"Successfully logged in user: {username}")
                flash('Login successful!', 'success')
                return redirect(url_for('dashboard'))
            else:
                logger.warning(f"Failed login attempt for username: {username}")
                flash('Invalid username or password', 'error')
                return render_template('login.html')

        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            flash('An error occurred during login', 'error')
            return render_template('login.html')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

@app.route('/add', methods=['POST'])
def add():
    if 'user_id' not in session:
        flash('Please log in to add entries', 'warning')
        return redirect(url_for('login'))
    
    try:
        date_str = request.form['date']
        hours = float(request.form['hours'])
        description = request.form.get('description', '').strip()

        # Input validation
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format', 'error')
            return redirect(url_for('dashboard'))

        if hours <= 0 or hours > 24:
            flash('Hours must be between 0 and 24', 'error')
            return redirect(url_for('dashboard'))

        entry = Entry(
            user_id=session['user_id'],
            date=date,
            hours=hours,
            description=description
        )
        db.session.add(entry)
        db.session.commit()
        
        flash('Entry added successfully!', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        logger.error(f"Error adding entry: {str(e)}", exc_info=True)
        db.session.rollback()
        flash('An error occurred while adding the entry', 'error')
        return redirect(url_for('dashboard'))

@app.route('/delete/<int:id>', methods=['POST'])
def delete(id):
    if 'user_id' not in session:
        flash('Please log in to delete entries', 'warning')
        return redirect(url_for('login'))
    
    try:
        entry = Entry.query.get_or_404(id)
        if entry.user_id != session['user_id']:
            flash('Unauthorized access', 'error')
            return redirect(url_for('dashboard'))

        db.session.delete(entry)
        db.session.commit()

        # Export to Excel after successful deletion
        if export_to_excel():
            flash('Entry deleted and backup updated successfully', 'success')
        else:
            flash('Entry deleted but Excel backup failed', 'warning')

        return redirect(url_for('dashboard'))

    except Exception as e:
        logger.error(f"Error deleting entry: {str(e)}")
        db.session.rollback()
        flash('An error occurred while deleting the entry', 'error')
        return redirect(url_for('dashboard'))

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if 'user_id' not in session:
        flash('Please log in to access settings', 'warning')
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])
    if not user.settings:
        user.settings = UserSettings()
        db.session.commit()
    
    if request.method == 'POST':
        try:
            user.settings.full_name = request.form.get('full_name', '').strip()
            user.settings.hourly_rate = float(request.form.get('hourly_rate', 0))
            user.settings.tax_rate = float(request.form.get('tax_rate', 0))
            db.session.commit()
            flash('Settings updated successfully', 'success')
            return redirect(url_for('settings'))
        except ValueError:
            flash('Invalid input values', 'error')
        except Exception as e:
            logger.error(f"Error updating settings: {str(e)}")
            flash('An error occurred while updating settings', 'error')
    
    return render_template('settings.html', settings=user.settings)

@app.route('/calculate_salary', methods=['POST'])
def calculate_salary():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    try:
        user = User.query.get(session['user_id'])
        if not user.settings or not user.settings.hourly_rate:
            return jsonify({'error': 'Please set your hourly rate in settings'}), 400
        
        hours = float(request.form.get('hours', 0))
        hourly_rate = user.settings.hourly_rate
        tax_rate = user.settings.tax_rate / 100  # Convert percentage to decimal
        
        gross_pay = hours * hourly_rate
        tax_amount = gross_pay * tax_rate
        net_pay = gross_pay - tax_amount
        
        return jsonify({
            'gross_pay': round(gross_pay, 2),
            'tax_amount': round(tax_amount, 2),
            'net_pay': round(net_pay, 2)
        })
    except Exception as e:
        logger.error(f"Error calculating salary: {str(e)}")
        return jsonify({'error': 'Error calculating salary'}), 500

@app.route('/check_hawaii_tax')
def check_hawaii_tax():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    try:
        user = User.query.get(session['user_id'])
        if not user.settings:
            user.settings = UserSettings()
        
        # Set Hawaii tax rate (approximately 11% for most income brackets)
        user.settings.tax_rate = 11.0
        db.session.commit()
        
        return jsonify({
            'message': 'Hawaii tax rate (11%) has been applied to your settings',
            'tax_rate': 11.0
        })
    except Exception as e:
        logger.error(f"Error setting Hawaii tax: {str(e)}")
        return jsonify({'error': 'Error setting Hawaii tax rate'}), 500

if __name__ == '__main__':
    app.run(debug=os.getenv('FLASK_DEBUG', '0') == '1')

from flask import session, redirect, request, abort, url_for

@app.route("/login/google")
def login_google():
    auth_url, state = flow.authorization_url()
    session["state"] = state
    return redirect(auth_url)

@app.route("/oauth2callback")
def oauth2callback():
    if request.args.get("state") != session.get("state"):
        abort(500, "Invalid state parameter")
    flow.fetch_token(authorization_response=request.url)

    creds = flow.credentials
    # now you can use creds.token, creds.id_token, etc.
    return redirect(url_for("dashboard"))
